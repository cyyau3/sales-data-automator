# src/web_navigator.py
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
from logger_config import logger
from datetime import datetime
import pandas as pd
import os
import time
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import traceback
from urls import URLConfig
from selenium.webdriver.chrome.options import Options
import subprocess
from pathlib import Path
import calendar


class WebNavigator:
    def __init__(self, timeout=30):
        """Initialize WebNavigator with directories setup"""
        self.timeout = timeout
        
        # Setup directories using Path
        self._project_root = Path(__file__).parent.parent
        self._exports_dir = self._project_root / 'exports'
        self._downloads_dir = self._exports_dir / 'downloads'
        
        # Create necessary directories
        self._exports_dir.mkdir(exist_ok=True)
        self._downloads_dir.mkdir(exist_ok=True)
        
        # Store string versions for JSON-serializable contexts
        self.project_root = str(self._project_root)
        self.exports_dir = str(self._exports_dir)
        self.downloads_dir = str(self._downloads_dir)
        
        # Add report configurations
        self.report_configs = {
            "sum_by_week": {
                "filename": "連鎖通路商品週銷售報表(依週期).xls",
                "sheet_name": "Weekly Summary"
            },
            "sum_by_week_customer": {
                "filename": "連鎖通路商品週銷售報表(依通路).xls",
                "sheet_name": "Weekly Customer Summary"
            },
            "sum_by_month": {
                "filename": "連鎖通路商品月銷售報表(依期間).xls",
                "sheet_name": "Monthly Summary"
            },
            "sum_by_month_customer": {
                "filename": "連鎖通路商品月銷售報表(依客戶).xls",
                "sheet_name": "Monthly Customer Summary"
            }
        }
        
        logger.info(f"Downloads directory set to: {self.downloads_dir}")
        
        try:
            chrome_options = Options()
            
            # Configure Chrome options for automatic downloads
            chrome_options.add_experimental_option('prefs', {
                'download.default_directory': self.downloads_dir,
                'download.prompt_for_download': False,
                'download.directory_upgrade': True,
                'safebrowsing.enabled': True
            })
            
            # Initialize Chrome WebDriver with options
            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.maximize_window()
            self.wait = WebDriverWait(self.driver, timeout)
            
        except Exception as e:
            logger.error(f"Failed to initialize WebNavigator: {str(e)}")
            raise

    def _get_downloads_path(self) -> Path:
        """Get downloads directory as Path object"""
        return Path(self.downloads_dir)

    def _get_exports_path(self) -> Path:
        """Get exports directory as Path object"""
        return Path(self.exports_dir)

    def login(self, username, password):
        """Login to UCD website"""
        try:
            # Navigate to main page
            logger.info("Navigating to main page...")
            self.driver.get(URLConfig.BASE_URL)
            
            # Wait for and click the login link in nav bar
            logger.info("Looking for login link...")
            login_link = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, f"//a[contains(@href, '{URLConfig.LOGIN_PATH}')]"))
            )
            logger.info("Clicking login link...")
            login_link.click()
            
            # Wait for login form
            logger.info("Waiting for login form...")
            
            # Find form elements using their exact IDs
            username_field = self.wait.until(
                EC.presence_of_element_located((By.ID, "user_name"))
            )
            password_field = self.wait.until(
                EC.presence_of_element_located((By.ID, "user_password"))
            )
            
            # Clear and fill in credentials
            username_field.clear()
            logger.info("Entering username...")
            username_field.send_keys(username)
            
            password_field.clear()
            logger.info("Entering password...")
            password_field.send_keys(password)
            
            # Click the login button using its exact attributes
            login_button = self.wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='B1'][type='submit'][value='確認登入']"))
            )
            logger.info("Clicking submit button...")
            login_button.click()
            
            # Wait for redirect to member page
            logger.info("Waiting for redirect to member page...")
            self.wait.until(
                EC.url_to_be(URLConfig.get_full_url(URLConfig.MEMBER_PATH))
            )
            
            logger.info("Successfully logged in")
                
        except TimeoutException as e:
            logger.error("Login form interaction failed - timeout")
            self.save_screenshot("login_failure")
            raise
        except Exception as e:
            # Mask the username in the error message
            logger.error(f"Login process failed for user: {username[:2]}***")
            self.save_screenshot("login_failure")
            raise

    def return_to_index(self):
        """Return to the member index page"""
        try:
            # Navigate directly to index page
            self.driver.get(URLConfig.get_full_url(URLConfig.MEMBER_PATH))
            
            # Additional wait for nav menu
            self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "nav"))
            )
            
            logger.info("Successfully returned to index page")
            
        except Exception as e:
            logger.error(f"Failed to return to index: {str(e)}")
            self.save_screenshot("return_to_index_error")
            raise

    def filter_month_generator(self, year=None, month=None):
        """Generate appropriate month and year for filtering"""
        try:
            current_date = datetime.now()
            
            # If no year/month provided, get previous month
            if year is None and month is None:
                if current_date.month == 1:
                    year = current_date.year - 1
                    month = 12
                else:
                    year = current_date.year
                    month = current_date.month - 1
            
            # Validate month
            if month < 1 or month > 12:
                raise ValueError(f"Invalid month value: {month}")
                
            # Return both separate and combined formats
            return {
                'year': year,
                'month': str(month).zfill(2),
                'combined': f"{year}{str(month).zfill(2)}"  # e.g., "202410"
            }
            
        except Exception as e:
            logger.error(f"Failed to generate filter dates: {str(e)}")
            raise

    def navigate_to_inventory(self):
        """Navigate to the inventory page after login"""
        try:
            # Use self.wait instead of creating new WebDriverWait
            nav_div = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "nav"))
            )
            
            inventory_link = self.driver.find_element(By.XPATH, "//a[contains(text(), '[606030] 庫存明細')]")
            inventory_link.click()
            
            # Use self.wait instead of creating new WebDriverWait
            self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "dataGrid"))
            )
            
            logger.info("Successfully navigated to inventory page")
            
        except TimeoutException:
            logger.error("Timeout waiting for inventory page elements")
            raise
        except Exception as e:
            logger.error(f"Failed to navigate to inventory page: {str(e)}")
            raise

    def extract_inventory_table(self):
        """Extract data from the inventory table"""
        try:
            # Wait for table to be present
            table = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "dataGrid"))
            )
            
            # Initialize lists to store data
            data = []
            
            # Get headers
            headers = []
            header_row = table.find_element(By.TAG_NAME, "thead").find_element(By.TAG_NAME, "tr")
            for th in header_row.find_elements(By.TAG_NAME, "th"):
                headers.append(th.text.strip())
            
            # Get table body rows
            tbody = table.find_element(By.TAG_NAME, "tbody")
            rows = tbody.find_elements(By.TAG_NAME, "tr")
            
            # Extract data from each row
            for row in rows:
                row_data = []
                for cell in row.find_elements(By.TAG_NAME, "td"):
                    row_data.append(cell.text.strip())
                if row_data:  # Only add non-empty rows
                    data.append(row_data)
            
            # Extract footer data
            tfoot = table.find_element(By.TAG_NAME, "tfoot")
            footer_row = tfoot.find_element(By.TAG_NAME, "tr")
            
            footer_data = []
            # Get text from pdtCode (總計)
            footer_data.append(footer_row.find_element(By.CLASS_NAME, "pdtCode").text.strip())
            
            # Get text and number from pdtName (共19種產品)
            pdt_name_text = footer_row.find_element(By.CLASS_NAME, "pdtName").text.strip()
            footer_data.append(pdt_name_text)
            
            # Get number from stockQuantity
            stock_qty = footer_row.find_element(By.CLASS_NAME, "stockQuantity").text.strip()
            footer_data.append(stock_qty)
            
            # Get number from stockAmount
            stock_amount = footer_row.find_element(By.CLASS_NAME, "stockAmount").text.strip()
            footer_data.append(stock_amount)
            
            # Add empty values for remaining columns (定價, 序號, 安全存量)
            footer_data.extend(['', '', ''])
            
            # Add footer data to main data
            data.append(footer_data)


            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Convert numeric columns
            numeric_columns = ['庫存量', '庫存額', '定價', '安全存量']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            logger.info(f"Successfully extracted {len(df)} inventory records")
            return df
            
        except Exception as e:
            logger.error(f"Failed to extract inventory table: {str(e)}")
            raise

    def navigate_to_monthly_supply(self):
            """Navigate to the monthly supply report page"""
            try:
                nav_div = self.wait.until(
                    EC.presence_of_element_located((By.CLASS_NAME, "nav"))
                )
                
                # Update this XPath according to the actual menu item text
                supply_link = self.driver.find_element(By.XPATH, "//a[contains(text(), '[606031] 庫存月報表')]")
                supply_link.click()
                
                # Wait for the filter form to be present
                self.wait.until(
                    EC.presence_of_element_located((By.XPATH, "//form[@action='supp_summary.jsp']"))
                )
                
                logger.info("Successfully navigated to monthly supply page")
                
            except TimeoutException:
                logger.error("Timeout waiting for monthly supply page elements")
                raise
            except Exception as e:
                logger.error(f"Failed to navigate to monthly supply page: {str(e)}")
                raise

    def set_monthly_supply_filter(self, year=None, month=None):
        """Set filter for monthly supply report"""
        try:
            # Get date values
            date_values = self.filter_month_generator(year, month)
            
            # Select year
            year_select = self.wait.until(
                EC.presence_of_element_located((By.NAME, "p_year"))
            )
            year_dropdown = Select(year_select)
            year_dropdown.select_by_value(str(date_values['year']))

            # Select month
            month_select = self.wait.until(
                EC.presence_of_element_located((By.NAME, "p_period"))
            )
            month_dropdown = Select(month_select)
            month_dropdown.select_by_value(date_values['month'])

            # Submit
            submit_button = self.wait.until(
                EC.element_to_be_clickable((By.NAME, "B1"))
            )
            submit_button.click()

            # Wait for results
            self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "sortable"))
            )
            
            logger.info(f"Successfully set filter for {date_values['year']}/{date_values['month']}")

        except Exception as e:
            logger.error(f"Failed to set monthly supply filter: {str(e)}")
            raise

    def extract_monthly_supply_table(self):
        """Extract data from the monthly supply table"""
        try:
            # Extract title from p element
            title = self.driver.find_element(By.XPATH, "//p[contains(text(), '庫存銷售月報表')]").text
            logger.debug(f"Found title: {title}")

            # Wait for main table to be present
            main_table = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "sortable"))
            )
            
            # Get the main table data
            table_html = main_table.get_attribute('outerHTML')
            tables = pd.read_html(table_html)
            df = tables[0]
            
            # Convert columns as before
            numeric_columns = ['定價', '存量', '存額', '月進量', '退量', '進淨量', 
                            '出量', '退量', '出淨量', '年量', '退量', '進淨量', 
                            '出量', '退量', '出淨量']
            
            date_columns = ['發書日']
            string_columns = ['貨物代碼', '書名', '系列編號']
            
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(',', '')
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            for col in date_columns:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
            
            for col in string_columns:
                if col in df.columns:
                    df[col] = df[col].fillna('').astype(str)

            try:
                # Try to find the summary row using a more specific XPath
                summary_rows = self.driver.find_elements(
                    By.XPATH, "//tr[td[contains(text(), '合計') or contains(text(), '合  計')]]"
                )
                
                if summary_rows:
                    logger.debug(f"Found {len(summary_rows)} potential summary rows")
                    summary_row = summary_rows[-1]  # Take the last one if multiple found
                    cells = summary_row.find_elements(By.TAG_NAME, "td")
                    
                    # Log the actual cell contents for debugging
                    cell_texts = [cell.text.strip() for cell in cells]
                    logger.debug(f"Summary row cell contents: {cell_texts}")
                    
                    # Create summary data dictionary with empty values for first columns
                    summary_data = {
                        '貨物代碼': '',
                        '書名': '',
                        '發書日': pd.NaT,
                        '定價': None,
                        '系列編號': '合計'  # Put '計' in 系列編號 column
                    }
                    
                    # Remove the first cell that contains '合計' and process the remaining cells sequentially
                    remaining_cells = cells[1:]  # Skip the first cell with '合計'
                    
                    # Map the values to the correct columns in order
                    columns_order = ['存量', '存額', '月進量', '退量', '進淨量', 
                                   '出量', '退量.1', '出淨量', '年進量', '退量.2', 
                                   '進淨量.1', '出量.1', '退量.3', '出淨量.1']
                    
                    for i, col_name in enumerate(columns_order):
                        if i < len(remaining_cells):
                            value = remaining_cells[i].text.strip().replace(',', '')
                            try:
                                summary_data[col_name] = float(value) if value else 0.0
                            except ValueError:
                                logger.warning(f"Could not convert value for {col_name}: {value}")
                                summary_data[col_name] = 0.0
                        else:
                            summary_data[col_name] = 0.0
                    
                    # Add summary row to DataFrame
                    summary_df = pd.DataFrame([summary_data])
                    df = pd.concat([df, summary_df], ignore_index=True)
                    logger.debug(f"Added summary row: {summary_data}")
                else:
                    logger.warning("No summary row found")

            except Exception as e:
                logger.warning(f"Failed to extract summary data: {str(e)}")
                logger.warning(f"Summary extraction error details: {traceback.format_exc()}")

            logger.info(f"Successfully extracted {len(df)} monthly supply records")
            return df, title
            
        except Exception as e:
            logger.error(f"Failed to extract monthly supply table: {str(e)}")
            raise

    def navigate_to_analysis_report(self):
        """Navigate to the analysis report page"""
        try:
            nav_div = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "nav"))
            )
            
            # Update XPath to match the exact menu text
            analysis_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(), '[606062] 銷售資料綜合分析')]")
                )
            )
            analysis_link.click()
            
            # Wait for the filter form to load
            self.wait.until(
                EC.presence_of_element_located((By.NAME, "b_ym"))
            )
            
            logger.debug("Analysis report page loaded")  # Add debug logging
            logger.info("Successfully navigated to analysis report page")
            
        except TimeoutException:
            logger.error("Timeout waiting for analysis report page elements")
            self.save_screenshot("analysis_navigation_error")
            raise
        except Exception as e:
            logger.error(f"Failed to navigate to analysis report page: {str(e)}")
            self.save_screenshot("analysis_navigation_error")
            raise
    
    def set_analysis_report_filter(self, year=None, month=None, filter_type='customer'):
        """Set filter for analysis report
        Args:
            year: Optional year to filter
            month: Optional month to filter
            filter_type: 'customer' or 'product' to determine which checkboxes to select
        """
        try:
            # Get date values
            date_values = self.filter_month_generator(year, month)
            combined_date = date_values['combined']
            
            logger.debug(f"Setting analysis filter for {combined_date}, type: {filter_type}")
            
            # Select start and end dates (same month for our case)
            for field_name in ['b_ym', 'e_ym']:
                date_select = self.wait.until(
                    EC.presence_of_element_located((By.NAME, field_name))
                )
                date_dropdown = Select(date_select)
                date_dropdown.select_by_value(combined_date)
            
            # Clear existing selections
            checkboxes = self.driver.find_elements(By.XPATH, "//input[@type='checkbox']")
            for checkbox in checkboxes:
                if checkbox.is_selected():
                    checkbox.click()
            
            # Select appropriate checkboxes based on filter type
            if filter_type == 'customer':
                # Wait and select customer-related checkboxes
                self.wait.until(
                    EC.element_to_be_clickable((By.NAME, "acc_code"))
                ).click()
                self.wait.until(
                    EC.element_to_be_clickable((By.NAME, "acc_cat1"))
                ).click()
            else:  # product
                # Wait and select product-related checkboxes
                self.wait.until(
                    EC.element_to_be_clickable((By.NAME, "stk_c"))
                ).click()
                self.wait.until(
                    EC.element_to_be_clickable((By.NAME, "acc_cat"))
                ).click()
            
            # Submit form
            submit_button = self.wait.until(
                EC.element_to_be_clickable((By.NAME, "B1"))
            )
            submit_button.click()
            
            # Wait for results table
            self.wait.until(
                EC.presence_of_element_located((By.XPATH, "//table[@bgcolor='#008080']"))
            )
            
            logger.info(f"Successfully set analysis filter type: {filter_type} for {combined_date}")

        except Exception as e:
            logger.error(f"Failed to set analysis report filter: {str(e)}")
            self.save_screenshot("analysis_filter_error")
            raise

    def extract_analysis_table(self):
        """Extract data from analysis report table based on current filter"""
        try:
            # Wait for table to be present
            table = self.wait.until(
                EC.presence_of_element_located((By.XPATH, "//table[@bgcolor='#008080']"))
            )
            
            # Get headers
            headers = []
            header_row = table.find_element(By.TAG_NAME, "tr")
            for th in header_row.find_elements(By.TAG_NAME, "td"):
                headers.append(th.text.strip())
            
            # Initialize lists to store data
            data = []
            
            # Get all rows except header
            rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip header row
            
            # Process regular rows
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                row_data = []
                
                # Check if this is the total row
                is_total = False
                if cells[0].get_attribute("bgcolor") == "#CCFF66":
                    is_total = True
                
                for cell in cells:
                    value = cell.text.strip()
                    # If it's a total row and the cell spans multiple columns
                    if is_total and cell.get_attribute("colspan"):
                        row_data.append("合計")  # Add '合計' for the first column
                        # Add empty strings for spanned columns
                        for _ in range(int(cell.get_attribute("colspan")) - 1):
                            row_data.append("")
                    else:
                        row_data.append(value)
                        
                if row_data:  # Only add non-empty rows
                    data.append(row_data)
            
            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Convert numeric columns
            numeric_columns = ['出量', '退量', '淨量']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col].replace('', '0'), errors='coerce')
            
            # Convert 退率 (return rate) - remove % and convert to numeric
            if '退率' in df.columns:
                df['退率'] = df['退率'].replace('', '0')
                df['退率'] = df['退率'].str.rstrip('%').astype(float)
            
            logger.info(f"Successfully extracted {len(df)} analysis records")
            return df
                
        except Exception as e:
            logger.error(f"Failed to extract analysis table: {str(e)}")
            self.save_screenshot("analysis_table_extraction_error")
            raise

    def export_to_excel(self, df, report_type, title=None, excel_path=None):
        """Export the DataFrame to Excel with report type specification and optional title
        Args:
            df: DataFrame to export
            report_type: Type of report (will be used as sheet name)
            title: Optional title for the sheet
            excel_path: Optional path to existing Excel file. If None, creates new file
        Returns:
            Path to the Excel file
        """
        try:
            # If no excel_path provided, create new file
            if excel_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"sales_data_{timestamp}.xlsx"
                
                current_dir = os.path.dirname(os.path.abspath(__file__))
                project_root = os.path.dirname(current_dir)
                exports_dir = os.path.join(project_root, 'exports')
                os.makedirs(exports_dir, exist_ok=True)
                
                excel_path = os.path.join(exports_dir, filename)
            
            # Check if file exists to determine mode
            mode = 'a' if os.path.exists(excel_path) else 'w'
            
            # Create Excel writer object with appropriate mode
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode) as writer:
                # Always write DataFrame, adjust startrow based on title presence
                start_row = 1 if title else 0
                df.to_excel(writer, sheet_name=report_type, index=False, startrow=start_row)
                
                if title:
                    # Get the worksheet and write title
                    worksheet = writer.sheets[report_type]
                    worksheet.cell(row=1, column=1, value=title)
            
            logger.info(f"Successfully exported {report_type} to sheet in {excel_path}")
            return excel_path
                
        except Exception as e:
            logger.error(f"Failed to export to Excel: {str(e)}")
            raise

    def save_screenshot(self, prefix):
        """Save screenshot on failure"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"error_screenshots/{prefix}_{timestamp}.png"
            os.makedirs("error_screenshots", exist_ok=True)
            self.driver.save_screenshot(filename)
            logger.info(f"Screenshot saved as {filename}")
        except Exception as e:
            logger.error(f"Failed to save screenshot: {str(e)}")

    def logout_and_quit(self):
        """Logout from the website and close the browser"""
        try:
            # Check if already logged out
            if not self.is_logged_in():
                logger.info("Already logged out")
                return

            # Find and click the logout link in nav bar
            logger.info("Looking for logout link...")
            logout_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, f"//a[contains(@href, '{URLConfig.LOGOUT_PATH}')][text()='會員登出']")
                )
            )
            logger.info("Clicking logout link...")
            logout_link.click()
            
            # Wait for logout message
            self.wait.until(
                EC.presence_of_element_located((By.XPATH, "//*[contains(text(), '您目前登出系統中')]"))
            )
            
            # Wait for redirect to homepage
            self.wait.until(
                lambda driver: driver.current_url == URLConfig.BASE_URL + "/index.jsp"
            )
            
            # Clear cookies and session storage
            self.driver.delete_all_cookies()
            self.driver.execute_script("window.localStorage.clear();")
            self.driver.execute_script("window.sessionStorage.clear();")
            
            logger.info("Successfully logged out")
        except Exception as e:
            logger.error(f"Logout failed: {str(e)}")
            raise  # Re-raise the exception to handle it at a higher level
        finally:
            self.close()

    def close(self):
        """Close the browser"""
        try:
            if self.driver:
                self.driver.quit()
                logger.info("Browser session closed")
        except Exception as e:
            logger.error(f"Error closing browser: {str(e)}")
        finally:
            self.driver = None

    def is_logged_in(self):
        """Check if user is currently logged in"""
        try:
            logout_link = self.driver.find_element(By.XPATH, "//a[contains(@href, '/user_menu/user_logout.jsp')][text()='會員登出']")
            return logout_link.is_displayed()
        except:
            return False

    def extract_analysis_table(self):
        """Extract data from analysis report table based on current filter"""
        try:
            # Wait for table to be present
            table = self.wait.until(
                EC.presence_of_element_located((By.XPATH, "//table[@bgcolor='#008080']"))
            )
            
            # Get headers
            headers = []
            header_row = table.find_element(By.TAG_NAME, "tr")
            for th in header_row.find_elements(By.TAG_NAME, "td"):
                headers.append(th.text.strip())
            
            # Initialize lists to store data
            data = []
            
            # Get all rows except header
            rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip header row
            
            # Process regular rows
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                row_data = []
                
                # Check if this is the total row
                is_total = False
                if cells[0].get_attribute("bgcolor") == "#CCFF66":
                    is_total = True
                
                for cell in cells:
                    value = cell.text.strip()
                    # If it's a total row and the cell spans multiple columns
                    if is_total and cell.get_attribute("colspan"):
                        row_data.append("合計")  # Add '合計' for the first column
                        # Add empty strings for spanned columns
                        for _ in range(int(cell.get_attribute("colspan")) - 1):
                            row_data.append("")
                    else:
                        row_data.append(value)
                        
                if row_data:  # Only add non-empty rows
                    data.append(row_data)
            
            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Convert numeric columns
            numeric_columns = ['出量', '退量', '淨量']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col].replace('', '0'), errors='coerce')
            
            # Convert 退率 (return rate) - remove % and convert to numeric
            if '退率' in df.columns:
                df['退率'] = df['退率'].replace('', '0')
                df['退率'] = df['退率'].str.rstrip('%').astype(float)
            
            logger.info(f"Successfully extracted {len(df)} analysis records")
            return df
                
        except Exception as e:
            logger.error(f"Failed to extract analysis table: {str(e)}")
            self.save_screenshot("analysis_table_extraction_error")
            raise

    def navigate_to_weekly_summary(self):
        """Navigate to the sum by week menu page"""
        try:
            # Wait for navigation menu
            nav_div = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "nav"))
            )
            
            # Find and click the sum by week link
            sum_by_week_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(), '[606066] 連鎖通路商品週銷售報表')]")
                )
            )
            sum_by_week_link.click()
            
            logger.info("Successfully navigated to sum by week menu page")
            
        except Exception as e:
            logger.error(f"Failed to navigate to sum by week menu: {str(e)}")
            self.save_screenshot("sum_by_week_navigation_error")
            raise

    def process_downloaded_excel(self, file_path):
        """Convert downloaded xls file to xlsx format using LibreOffice
        Args:
            file_path: Path to the downloaded xls file
        Returns:
            Path to the converted xlsx file
        """
        try:
            # Security check: Validate file path
            file_path = Path(file_path).resolve()
            if not file_path.exists():
                raise FileNotFoundError(f"Input file not found: {file_path}")
            
            # Security check: Ensure file is within allowed directory
            if not str(file_path).startswith(str(Path(self.downloads_dir).resolve())):
                raise SecurityError("File path is outside allowed directory")
            
            input_path = str(file_path)
            output_dir = str(file_path.parent)
            
            logger.info(f"Input path: {input_path}")
            
            # List files before conversion
            files_before = os.listdir(output_dir)
            logger.info(f"Files before conversion: {files_before}")
            
            # Build and execute conversion command
            command = [
                '/Applications/LibreOffice.app/Contents/MacOS/soffice',
                '--headless',
                '--norestore',
                '--nofirststartwizard',
                '--nologo',
                '--convert-to', 'xlsx:Calc MS Excel 2007 XML',
                '--outdir', output_dir,
                input_path
            ]
            
            # For debugging only
            # logger.debug(f"Running command: {' '.join(command)}")
            
            process = subprocess.run(
                command,
                capture_output=True,
                text=True,
                check=True
            )
            
            # For debugging only
            # logger.debug(f"Command stdout: {process.stdout}")
            # if process.stderr:
            #     logger.warning(f"Command stderr: {process.stderr}")
            
            # List files after conversion
            files_after = os.listdir(output_dir)
            logger.info(f"Files after conversion: {files_after}")
            
            # Determine output path
            output_path = file_path.with_suffix('.xlsx')
            # logger.debug(f"Expected output path: {output_path}")
            
            # Verify conversion success
            if output_path.exists():
                logger.info("Conversion successful!")
                logger.info(f"Output file exists at: {output_path}")
                
                # Security check: Verify file size
                if output_path.stat().st_size == 0:
                    raise SecurityError("Converted file is empty")
                
                # Cleanup original file
                file_path.unlink()
                return output_path
            else:
                raise FileNotFoundError("Conversion failed - output file not found")
            
        except subprocess.CalledProcessError as e:
            logger.error(f"Conversion failed with return code {e.returncode}")
            if e.stdout:
                logger.error(f"stdout: {e.stdout}")
            if e.stderr:
                logger.error(f"stderr: {e.stderr}")
            raise
        except Exception as e:
            logger.error(f"Failed to process Excel file: {str(e)}")
            raise

    def navigate_to_monthly_summary(self):
        """Navigate to the monthly summary page"""
        try:
            # Wait for navigation menu
            nav_div = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "nav"))
            )
            
            # Find and click the monthly summary link
            monthly_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(), '[606067] 連鎖通路商品月銷售報表')]")
                )
            )
            monthly_link.click()
            
            logger.info("Successfully navigated to monthly summary page")
            
        except Exception as e:
            logger.error(f"Failed to navigate to monthly summary menu: {str(e)}")
            self.save_screenshot("monthly_summary_navigation_error")
            raise

    def set_report_filter(self, report_type):
        """Generic filter setter for both weekly and monthly reports"""
        try:
            # Get month value using your existing generator
            date_values = self.filter_month_generator()
            target_month = date_values['combined']
            logger.debug(f"Filtering for {date_values['year']}/{date_values['month']}")

            # Select the appropriate form based on report type
            form_links = {
                "sum_by_week": "//a[contains(text(), '連鎖通路商品週銷售報表(依週期)')]",
                "sum_by_week_customer": "//a[contains(text(), '連鎖通路商品週銷售報表(依客戶)')]",
                "sum_by_month": "//a[contains(text(), '連鎖通路商品月銷售報表(依期間)')]",
                "sum_by_month_customer": "//a[contains(text(), '連鎖通路商品月銷售報表(依客戶)')]"
            }

            form_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, form_links[report_type])
                )
            )
            form_link.click()

            # Handle different filter fields based on report type
            if report_type.startswith('sum_by_week'):
                # Weekly report filter fields
                start_select = Select(self.wait.until(
                    EC.presence_of_element_located((By.NAME, "mas_date_b"))
                ))
                end_select = Select(self.wait.until(
                    EC.presence_of_element_located((By.NAME, "mas_date_e"))
                ))

                # Get all options
                start_options = [opt.get_attribute('value') for opt in start_select.options]
                end_options = [opt.get_attribute('value') for opt in end_select.options]

                # Filter options for the target month
                month_start_options = [opt for opt in start_options if opt.startswith(target_month)]
                month_end_options = [opt for opt in end_options if opt.startswith(target_month)]

                if not month_start_options or not month_end_options:
                    raise ValueError(f"No options found for {date_values['year']}/{date_values['month']}")

                # Select first and last options for the month
                start_select.select_by_value(month_start_options[0])
                end_select.select_by_value(month_end_options[-1])

            else:
                # Monthly report filter fields
                start_field = self.wait.until(
                    EC.presence_of_element_located((By.NAME, "ym_b"))
                )
                end_field = self.wait.until(
                    EC.presence_of_element_located((By.NAME, "ym_e"))
                )
                Select(start_field).select_by_value(target_month)
                Select(end_field).select_by_value(target_month)

            # Submit the form
            submit_button = self.wait.until(
                EC.element_to_be_clickable((By.NAME, "B1"))  # Changed to match your original
            )
            submit_button.click()

            logger.info(f"Successfully set filter for {report_type} report: {date_values['year']}/{date_values['month']}")

        except Exception as e:
            logger.error(f"Failed to set filter for {report_type}: {str(e)}")
            self.save_screenshot(f"{report_type}_filter_error")
            raise

    def process_summary_reports(self, excel_path, report_category):
        """Process both weekly and monthly summary reports"""
        try:
            # Define report pairs
            report_pairs = {
                'weekly': {
                    'menu_func': self.navigate_to_weekly_summary,
                    'reports': ['sum_by_week', 'sum_by_week_customer']
                },
                'monthly': {
                    'menu_func': self.navigate_to_monthly_summary,
                    'reports': ['sum_by_month', 'sum_by_month_customer']
                }
            }

            pair = report_pairs[report_category]
            converted_files = []  # Track converted files
            
            for report_type in pair['reports']:
                # Navigate to appropriate menu
                pair['menu_func']()
                self.set_report_filter(report_type)
                
                # Wait for download
                file_path = os.path.join(self.downloads_dir, self.report_configs[report_type]["filename"])
                wait_start = time.time()
                while not os.path.exists(file_path) and time.time() - wait_start < 30:
                    time.sleep(0.5)
                    
                if not os.path.exists(file_path):
                    raise FileNotFoundError(f"Download timeout: {self.report_configs[report_type]['filename']}")
                
                # Convert file and store path
                xlsx_path = self.process_downloaded_excel(file_path)
                converted_files.append({
                    'path': xlsx_path,
                    'type': report_type,
                    'config': self.report_configs[report_type]
                })
                
                # Return to index for next report
                self.return_to_index()
                
            # Now process all converted files
            for file_info in converted_files:
                try:
                    # Read the converted file
                    df = pd.read_excel(file_info['path'], engine='openpyxl')
                    
                    # Get the header value from the first cell
                    header_value = df.columns[0]
                    
                    # Reset the column names to be blank after the first column
                    new_columns = [header_value] + [''] * (len(df.columns) - 1)
                    df.columns = new_columns

                    # Append to main report with merged header
                    with pd.ExcelWriter(str(excel_path), engine='openpyxl', mode='a') as writer:
                        if file_info['config']["sheet_name"] in writer.book.sheetnames:
                            idx = writer.book.sheetnames.index(file_info['config']["sheet_name"])
                            writer.book.remove(writer.book.worksheets[idx])
                        
                        # Write the DataFrame
                        df.to_excel(
                            writer,
                            sheet_name=file_info['config']["sheet_name"],
                            index=False
                        )
                        
                        # Get the worksheet
                        worksheet = writer.book[file_info['config']["sheet_name"]]
                        
                        # Merge the header cells
                        worksheet.merge_cells(
                            start_row=1,
                            start_column=1,
                            end_row=1,
                            end_column=len(df.columns)
                        )
                        
                        # Set alignment for merged cell
                        merged_cell = worksheet.cell(row=1, column=1)
                        merged_cell.alignment = openpyxl.styles.Alignment(
                            horizontal='center',
                            vertical='center'
                        )

                    logger.info(f"Successfully appended {file_info['config']['filename']} to main report")

                except Exception as e:
                    logger.error(f"Failed to append {file_info['config']['filename']}: {str(e)}")
                    raise
                finally:
                    # Cleanup converted file
                    try:
                        if os.path.exists(file_info['path']):
                            os.remove(file_info['path'])
                    except Exception as e:
                        logger.warning(f"Could not remove temporary file: {e}")

            logger.info(f"Successfully processed {report_category} reports")
            return excel_path

        except Exception as e:
            logger.error(f"Failed to process {report_category} reports: {str(e)}")
            self.save_screenshot(f"{report_category}_reports_error")
            raise

    def navigate_to_orders(self):
        """Navigate to the order page"""
        try:
            # Wait for navigation menu
            nav_div = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "nav"))
            )
            
            # Find and click the monthly summary link
            monthly_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(), '[606072] 交易單據資料下載')]")
                )
            )
            monthly_link.click()
            
            logger.info("Successfully navigated to order page")
            
        except Exception as e:
            logger.error(f"Failed to navigate to order menu: {str(e)}")
            self.save_screenshot("order_navigation_error")
            raise

    def set_order_filter(self, order_type):
        """Set filter for order reports
        Args:
            order_type: 'GR' for purchase order or 'RNS' for return order
        """
        try:
            # Get date values for previous month
            date_values = self.filter_month_generator()
            year = date_values['year']
            month = date_values['month'].zfill(2)
            
            # Calculate last day of month
            last_day = calendar.monthrange(int(year), int(month))[1]
            
            # Format dates (DD-MM-YYYY)
            start_date = f"01-{month}-{year}"
            end_date = f"{last_day:02d}-{month}-{year}"
            
            # Select order type
            order_select = Select(self.wait.until(
                EC.presence_of_element_located((By.NAME, "mas_code"))
            ))
            order_select.select_by_value(order_type)
            
            # Set start date
            start_date_field = self.wait.until(
                EC.presence_of_element_located((By.NAME, "date1"))
            )
            # Clear any existing value and set new date using JavaScript
            self.driver.execute_script(
                f"arguments[0].value = '{start_date}';", 
                start_date_field
            )
            
            # Set end date
            end_date_field = self.wait.until(
                EC.presence_of_element_located((By.NAME, "date2"))
            )
            self.driver.execute_script(
                f"arguments[0].value = '{end_date}';", 
                end_date_field
            )
            
            # Submit form
            submit_button = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@value='送出查詢']"))
            )
            submit_button.click()
            
            logger.info(f"Successfully set filter for {order_type} orders: {start_date} to {end_date}")
            
        except Exception as e:
            logger.error(f"Failed to set order filter for {order_type}: {str(e)}")
            self.save_screenshot(f"order_filter_error_{order_type}")
            raise
    
    def extract_order_data(self, order_type):
        """Extract data from order report
        Args:
            order_type: 'GR' for purchase order or 'RNS' for return order
        """
        try:
            # Wait for table to be present
            table = self.wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//table[@border='0' and @width='100%']")
                )
            )
            
            # Get metadata from first row
            metadata_cell = table.find_element(By.XPATH, ".//tr[1]/td").text.strip()
            metadata_lines = metadata_cell.split('\n')
            order_type_text = metadata_lines[0].strip()  # "單別：GR"
            date_range_text = metadata_lines[1].strip()  # "日期：01-10-2024 至 31-10-2024"
            
            # Get column headers
            header_cells = table.find_elements(By.XPATH, ".//tr[2]/td")
            headers = [cell.text.strip() for cell in header_cells]
            num_columns = len(headers)
            
            # Create all rows in the correct order
            all_rows = [
                # First row: Order type (padded with empty strings)
                [order_type_text] + [''] * (num_columns - 1),
                # Second row: Date range (padded with empty strings)
                [date_range_text] + [''] * (num_columns - 1),
                # Third row: Column headers
                headers
            ]
            
            # Add data rows
            for row in table.find_elements(By.XPATH, ".//tr[position()>2]"):
                cells = row.find_elements(By.TAG_NAME, "td")
                if cells and len(cells) == num_columns:
                    # Clean up the last cell (publisher) by removing hidden inputs
                    publisher = cells[-1].text.strip()
                    
                    # Get all other cell values
                    row_data = [cell.text.strip() for cell in cells[:-1]]
                    row_data.append(publisher)
                    
                    if any(row_data):  # Only add non-empty rows
                        all_rows.append(row_data)
            
            # Create DataFrame directly from all rows
            df = pd.DataFrame(all_rows)
            
            logger.info(f"Successfully extracted {len(all_rows)-3} {order_type} order records")
            return df
            
        except Exception as e:
            logger.error(f"Failed to extract {order_type} order data: {str(e)}")
            self.save_screenshot(f"order_extract_error_{order_type}")
            raise

    def process_order_reports(self, excel_path):
        """Process both purchase and return order reports"""
        try:
            order_configs = {
                'GR': {
                    'sheet_name': 'Purchase Orders',
                    'description': 'purchase'
                },
                'RNS': {
                    'sheet_name': 'Return Orders',
                    'description': 'return'
                }
            }
            
            for order_type, config in order_configs.items():
                try:
                    # Navigate to orders page
                    self.navigate_to_orders()
                    
                    # Set filter and get data
                    self.set_order_filter(order_type)
                    df = self.extract_order_data(order_type)
                    
                    # First remove the numeric row if it exists
                    if df.iloc[0].astype(str).str.match(r'^\d+$').all():
                        df = df.iloc[1:].reset_index(drop=True)
                    
                    # Then add the title row
                    title_df = pd.DataFrame([[config['sheet_name']] + [''] * (len(df.columns) - 1)], columns=df.columns)
                    df = pd.concat([title_df, df], ignore_index=True)
                    
                    # Export to Excel
                    with pd.ExcelWriter(str(excel_path), engine='openpyxl', mode='a') as writer:
                        if config['sheet_name'] in writer.book.sheetnames:
                            idx = writer.book.sheetnames.index(config['sheet_name'])
                            writer.book.remove(writer.book.worksheets[idx])
                        
                        # Write the DataFrame
                        df.to_excel(
                            writer,
                            sheet_name=config['sheet_name'],
                            index=False,
                            header=False  # Don't write the numeric headers
                        )
                        
                        # Get the worksheet
                        worksheet = writer.book[config['sheet_name']]
                        
                        # Merge the title cells in the first row
                        worksheet.merge_cells(
                            start_row=1,
                            start_column=1,
                            end_row=1,
                            end_column=len(df.columns)
                        )
                        
                        # Style the merged cell
                        merged_cell = worksheet.cell(row=1, column=1)
                        merged_cell.alignment = openpyxl.styles.Alignment(
                            horizontal='center',
                            vertical='center'
                        )
                    
                    logger.info(f"Successfully exported {config['description']} orders to sheet in {excel_path}")
                    
                    # Return to index for next report
                    self.return_to_index()
                    
                except Exception as e:
                    logger.error(f"Failed to process {config['description']} orders: {str(e)}")
                    raise
            
            return excel_path
            
        except Exception as e:
            logger.error(f"Failed to process order reports: {str(e)}")
            self.save_screenshot("order_reports_error")
            raise
    
    def navigate_to_payment_menu(self):
        """Navigate to the payment menu"""
        try:
            # Wait for navigation menu
            nav_div = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "nav"))
            )
            
            # Find and click the payment menu link
            payment_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(), '[606076] 供應商對帳作業')]")
                )
            )
            payment_link.click()  # Fixed typo in variable name
            
            # Add wait for page load like other navigation functions
            self.wait.until(
                EC.presence_of_element_located((By.TAG_NAME, "form"))
            )
            
            logger.info("Successfully navigated to payment menu")
            
        except TimeoutException:  # Add specific timeout handling like other functions
            logger.error("Timeout waiting for payment menu elements")
            self.save_screenshot("payment_menu_navigation_error")
            raise
        except Exception as e:
            logger.error(f"Failed to navigate to payment menu: {str(e)}")
            self.save_screenshot("payment_menu_navigation_error")
            raise

    def navigate_to_discount_detail(self):
        """Navigate to the discount detail page"""
        try:
            # Find and click the discount detail link
            discount_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(), '[折讓明細(輸出檔)]')]")
                )
            )
            discount_link.click()
            
            # Wait for the filter form to be present
            self.wait.until(
                EC.presence_of_element_located((By.XPATH, "//input[@name='period']"))
            )
            
            logger.info("Successfully navigated to discount details page")
            
        except TimeoutException:
            logger.error("Timeout waiting for discount detail page elements")
            self.save_screenshot("discount_detail_navigation_error")
            raise
        except Exception as e:
            logger.error(f"Failed to navigate to discount detail: {str(e)}")
            self.save_screenshot("discount_detail_navigation_error")
            raise

    def set_discount_filter(self):
        """Set filter for discount detail report"""
        try:
            # Get date 2 months ago using filter_month_generator
            current_date = datetime.now()
            if current_date.month <= 2:
                year = current_date.year - 1
                month = current_date.month + 10  # If month is 1 or 2, go back to previous year
            else:
                year = current_date.year
                month = current_date.month - 2
                
            # Format as YYYYMM
            period = f"{year}{month:02d}"
            
            # Find and fill the period input
            period_input = self.wait.until(
                EC.presence_of_element_located((By.NAME, "period"))
            )
            period_input.clear()
            period_input.send_keys(period)
            
            # Click submit button
            submit_button = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@value='查詢'][@name='B1']"))
            )
            submit_button.click()
            
            logger.info(f"Successfully set discount filter for period: {period}")
            
        except TimeoutException:
            logger.error("Timeout waiting for discount filter elements")
            self.save_screenshot("discount_filter_error")
            raise
        except Exception as e:
            logger.error(f"Failed to set discount filter: {str(e)}")
            self.save_screenshot("discount_filter_error")
            raise

    def extract_discount_table(self):
        """Extract data from the discount detail table"""
        try:
            logger.debug("Starting extract_discount_table function")
            time.sleep(3)  # Wait for new tab to open
            
            # Get all window handles and switch to the new tab
            handles = self.driver.window_handles
            logger.debug(f"Found {len(handles)} browser windows")
            
            if len(handles) < 2:
                raise Exception("Expected new tab to open after filter submission, but no new tab found")
            
            # Switch to the new tab (last opened)
            self.driver.switch_to.window(handles[-1])
            logger.debug(f"Switched to new tab with URL: {self.driver.current_url}")
            
            # Find all tables in the new tab
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            logger.debug(f"Found {len(tables)} tables on page")
            
            if len(tables) >= 2:
                table = tables[1]  # Use second table
            else:
                raise Exception(f"Not enough tables found in results tab. Found: {len(tables)}")
                
            # Get headers
            headers = []
            header_cells = table.find_elements(By.XPATH, ".//tr[1]/td")
            for cell in header_cells:
                header_text = cell.text.strip()
                headers.append(header_text)
            logger.debug(f"Found headers: {headers}")
            
            # Get all rows
            all_rows = table.find_elements(By.TAG_NAME, "tr")
            logger.debug(f"Total rows found: {len(all_rows)}")
            
            # Process data rows (skip header)
            data = []
            for i, row in enumerate(all_rows[1:], 1):  # Skip header row
                cells = row.find_elements(By.TAG_NAME, "td")
                logger.debug(f"Row {i} has {len(cells)} cells")
                
                # Check if this is the total row
                if len(cells) == 2:  # Total row has 2 cells
                    total_text = cells[0].text.strip()
                    total_amount = cells[1].text.strip()
                    logger.debug(f"Found total row: {total_text}, {total_amount}")
                    continue
                
                # Regular data row
                row_data = []
                for j, cell in enumerate(cells):
                    value = cell.text.strip()
                    if j == 0:  # Date column
                        # Extract just the date part if it contains time
                        if "00:00:00" in value:
                            value = value.split()[0]  # Take only the date part
                    row_data.append(value)
                    
                if len(row_data) == len(headers):
                    data.append(row_data)
                    logger.debug(f"Added data row: {row_data}")
                else:
                    logger.warning(f"Skipping row with incorrect number of cells. Expected {len(headers)}, got {len(row_data)}")
            
            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)
            logger.debug(f"Created DataFrame with shape: {df.shape}")
            
            # Convert date column - now using the exact format from the table
            df['日期'] = pd.to_datetime(df['日期'], format='%Y/%m/%d').dt.date
            logger.debug(f"Date column after conversion:\n{df['日期']}")
            
            # Convert amount column
            df['折讓金額'] = df['折讓金額'].str.replace(',', '').astype(float)
            
            # Add total row
            total_data = {
                '日期': None,
                '折讓類別': None,
                '說明': total_text if 'total_text' in locals() else "合計",
                '折讓金額': float(total_amount.replace(',', '')) if 'total_amount' in locals() else 0.0
            }
            df = pd.concat([df, pd.DataFrame([total_data])], ignore_index=True)
            
            logger.info(f"Successfully extracted {len(df)-1} discount records plus total")
            logger.debug(f"Final DataFrame:\n{df}")
            
            # Before closing the tab, collect all URLs and their categories
            discount_links = []
            for row in all_rows[1:-1]:  # Skip header and total rows
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) == 4:  # Regular data row
                    cell = cells[2]  # The third column contains links
                    links = cell.find_elements(By.TAG_NAME, "a")
                    if links:
                        link = links[0]
                        category = link.text.strip()
                        url = link.get_attribute('href')
                        discount_links.append({
                            'category': category,
                            'url': url
                        })
            
            logger.debug(f"Found {len(discount_links)} discount detail links")
            
            # Process each link
            for link_data in discount_links:
                try:
                    # Store current window handle
                    current_window = self.driver.current_window_handle
                    
                    # Click the link (opens in new tab)
                    self.driver.execute_script(f"window.open('{link_data['url']}', '_blank');")
                    logger.debug(f"Triggered download for {link_data['category']}")
                    
                    # Don't switch to new tab, just wait for download
                    time.sleep(3)  # Wait for download to start
                    download_path = None
                    for _ in range(30):  # Wait up to 30 seconds
                        files = list(self._get_downloads_path().glob("*.xls"))
                        if files:
                            download_path = str(max(files, key=os.path.getctime))
                            break
                        time.sleep(1)
                    
                    if download_path:
                        logger.debug(f"Found downloaded file: {download_path}")
                        
                        # Read the Excel file with explicit engine specification
                        detail_df = pd.read_excel(download_path, engine='xlrd')
                        
                        # Append to main Excel file
                        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a') as writer:
                            if link_data['category'] in writer.book.sheetnames:
                                idx = writer.book.sheetnames.index(link_data['category'])
                                writer.book.remove(writer.book.worksheets[idx])
                            
                            detail_df.to_excel(
                                writer,
                                sheet_name=link_data['category'],
                                index=False
                            )
                            logger.info(f"Added detail sheet: {link_data['category']}")
                        
                        # Clean up downloaded file
                        try:
                            os.remove(download_path)
                            logger.debug(f"Cleaned up downloaded file: {download_path}")
                        except Exception as e:
                            logger.warning(f"Could not remove downloaded file: {e}")
                    
                    # Make sure we're on the correct window
                    if self.driver.current_window_handle != current_window:
                        self.driver.switch_to.window(current_window)
                    
                except Exception as e:
                    # Log as debug instead of error for expected empty files
                    logger.debug(f"Skipping discount detail link {link_data['category']}")
                    continue
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to extract discount details: {str(e)}\nTraceback: {traceback.format_exc()}")
            self.save_screenshot("discount_extract_error")
            raise
        finally:
            # Clean up: close the results tab and switch back to original
            if len(self.driver.window_handles) > 1:
                self.driver.close()
                self.driver.switch_to.window(handles[0])
                logger.debug("Switched back to original tab")

    def process_discount_report(self, excel_path):
        """Process discount report and export to Excel"""
        try:
            # Get main discount table
            df = self.extract_discount_table()
            
            # Export main discount table to Excel
            with pd.ExcelWriter(str(excel_path), engine='openpyxl', mode='a') as writer:
                sheet_name = 'Discount Details'
                
                # Remove sheet if it exists
                if sheet_name in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index(sheet_name)
                    writer.book.remove(writer.book.worksheets[idx])
                
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )
                
                # Format the worksheet
                worksheet = writer.book[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Process downloaded detail files
            downloads_path = self._get_downloads_path()
            for file in downloads_path.glob("*.xls"):
                try:
                    # Convert using LibreOffice
                    converted_path = self.process_downloaded_excel(file)
                    if converted_path:
                        # Read the Excel file with header=None to handle merged cells
                        df = pd.read_excel(converted_path, header=None)
                        
                        # Check if first row contains merged cells
                        first_row = df.iloc[0]
                        has_merged_header = first_row.isna().any()
                        
                        if has_merged_header:
                            # Get the first non-null value from the merged header
                            header_title = first_row.dropna().iloc[0]
                            logger.debug(f"Found merged header: {header_title}")
                            
                            # Skip the merged header row and use second row as column names
                            df = pd.read_excel(converted_path, header=1)
                        else:
                            # No merged cells, read normally
                            df = pd.read_excel(converted_path)
                        
                        # Get category name from filename
                        category = os.path.splitext(file.name)[0].replace("discount_", "")
                        sheet_name = f"Discount_{category}"
                        
                        # Append to main Excel
                        with pd.ExcelWriter(str(excel_path), engine='openpyxl', mode='a') as writer:
                            if sheet_name in writer.book.sheetnames:
                                idx = writer.book.sheetnames.index(sheet_name)
                                writer.book.remove(writer.book.worksheets[idx])
                            
                            df.to_excel(
                                writer,
                                sheet_name=sheet_name,
                                index=False
                            )
                            
                            # If there was a merged header, add it back with bold formatting
                            if has_merged_header:
                                worksheet = writer.book[sheet_name]
                                # Merge cells in first row
                                worksheet.insert_rows(0)
                                first_cell = worksheet.cell(row=1, column=1)
                                first_cell.value = header_title
                                
                                # Apply bold formatting to merged header
                                first_cell.font = Font(bold=True)
                                
                                # Merge cells and center align
                                worksheet.merge_cells(start_row=1, start_column=1, 
                                                   end_row=1, end_column=len(df.columns))
                                first_cell.alignment = Alignment(horizontal='center')
                        
                            logger.info(f"Added discount detail sheet: {sheet_name}")
                        
                            # Cleanup converted file
                            os.remove(converted_path)
                    
                except Exception as e:
                    logger.error(f"Failed to process detail file {file}: {str(e)}")
                    continue
                finally:
                    # Cleanup original file
                    if file.exists():
                        file.unlink()
                    
            logger.info(f"Successfully exported discount details to {excel_path}")
            return excel_path
        
        except Exception as e:
            logger.error(f"Failed to process discount report: {str(e)}")
            self.save_screenshot("discount_report_error")
            raise

    def navigate_to_payment_detail(self):
        """Navigate to the payment detail page"""
        try:
            # Find and click the payment detail link
            payment_link = self.wait.until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//a[contains(text(), '[付款明細]')]")
                )
            )
            payment_link.click()
            
            # Wait for the filter form to be present
            self.wait.until(
                EC.presence_of_element_located((By.NAME, "date1"))
            )
            
            logger.info("Successfully navigated to payment details page")
            
        except TimeoutException:
            logger.error("Timeout waiting for payment detail page elements")
            self.save_screenshot("payment_detail_navigation_error")
            raise
        except Exception as e:
            logger.error(f"Failed to navigate to payment detail: {str(e)}")
            self.save_screenshot("payment_detail_navigation_error")
            raise   

    def set_payment_filter(self):
        """Set date filter for payment detail report"""
        try:
            # Get date values for previous month
            date_values = self.filter_month_generator()
            year = date_values['year']
            month = int(date_values['month'])
            
            # Get the last day of the month
            last_day = calendar.monthrange(year, month)[1]
            
            logger.debug(f"Setting date range for year: {year}, month: {month}")
            
            # Set start date (1st of month)
            # Click first calendar icon
            start_calendar = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, 
                    "//input[@name='date1']/following-sibling::a/img[@src='date.gif']"
                ))
            )
            start_calendar.click()
            
            # Click the previous month arrow
            prev_month = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//img[@src='previ.gif']"))
            )
            prev_month.click()
            
            # Select day 1
            start_date = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, 
                    "//a[contains(@href, 'javascript:yxPickDate(1)')]/span[text()='1']"
                ))
            )
            start_date.click()
            
            # Small delay between date selections
            time.sleep(1)
            
            # Set end date (last day of month)
            # Click second calendar icon
            end_calendar = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, 
                    "//input[@name='date2']/following-sibling::a/img[@src='date.gif']"
                ))
            )
            end_calendar.click()
            
            # Click the previous month arrow again
            prev_month = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//img[@src='previ.gif']"))
            )
            prev_month.click()
            
            # Select the last day
            end_date = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, 
                    f"//a[contains(@href, 'javascript:yxPickDate({last_day})')]/span[text()='{last_day}']"
                ))
            )
            end_date.click()
            
            # Click submit button
            submit_button = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//input[@value='確定'][@name='B1']"))
            )
            submit_button.click()
            
            logger.info(f"Successfully set payment filter for period: 1-{month:02d}-{year} to {last_day}-{month:02d}-{year}")
            
        except TimeoutException:
            logger.error("Timeout waiting for payment filter elements")
            self.save_screenshot("payment_filter_error")
            raise
        except Exception as e:
            logger.error(f"Failed to set payment filter: {str(e)}")
            self.save_screenshot("payment_filter_error")
            raise

    def extract_payment_table_data(self, table_index=1, sheet_name="Table Data"):
        """Generic function to extract table data from a page
        Args:
            table_index: Which table to extract (0-based index)
            sheet_name: Name of the sheet for Excel export
        Returns:
            DataFrame if data exists, None if no data found
        """
        try:
            # Wait for new tab to open
            time.sleep(3)
            
            # Get all window handles and switch to the new tab
            handles = self.driver.window_handles
            logger.debug(f"Found {len(handles)} browser windows")
            
            if len(handles) < 2:
                # No new tab means no data
                logger.info("No payment data found for the selected period")
                return None
            
            # Switch to the new tab (last opened)
            self.driver.switch_to.window(handles[-1])
            logger.debug(f"Switched to new tab with URL: {self.driver.current_url}")
            
            # Find all tables in the new tab
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            logger.debug(f"Found {len(tables)} tables on page")
            
            if len(tables) < 2:
                logger.info("No payment data table found")
                self.driver.close()
                self.driver.switch_to.window(handles[0])
                return None
            
            target_table = tables[table_index]  # Use second table
            
            # Get headers
            headers = []
            header_row = target_table.find_element(By.TAG_NAME, "tr")
            for th in header_row.find_elements(By.TAG_NAME, "td"):
                headers.append(th.text.strip())
            
            # Get all rows except header
            rows = target_table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip header row
            
            # Check if there are any data rows
            if not rows:
                logger.info("No payment records found (empty table)")
                self.driver.close()
                self.driver.switch_to.window(handles[0])
                return None
                
            # Process data rows
            data = []
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                row_data = [cell.text.strip() for cell in cells]
                if any(row_data):  # Only add non-empty rows
                    data.append(row_data)
            
            # If no valid data rows were found
            if not data:
                logger.info("No payment records found (no valid data rows)")
                self.driver.close()
                self.driver.switch_to.window(handles[0])
                return None
                
            # Create DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # Convert numeric columns (金額)
            if '金額' in df.columns:
                df['金額'] = df['金額'].str.replace(',', '').astype(float)
            
            # Convert date columns (日期, 到期日) - handle YYYYMMDD format
            date_columns = ['日期', '到期日']
            for col in date_columns:
                if col in df.columns:
                    # First try YYYYMMDD format
                    df[col] = pd.to_datetime(df[col], format='%Y%m%d', errors='coerce')
                    # If any dates failed to parse, try YYYY/MM/DD as fallback
                    mask = df[col].isna()
                    if mask.any():
                        df.loc[mask, col] = pd.to_datetime(
                            df.loc[mask, col], 
                            format='%Y/%m/%d', 
                            errors='coerce'
                        )
            
            logger.info(f"Successfully extracted {len(df)} payment records")
            
            # Close the tab and switch back
            self.driver.close()
            self.driver.switch_to.window(handles[0])
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to extract table data: {str(e)}\nTraceback: {traceback.format_exc()}")
            self.save_screenshot("table_extract_error")
            # Make sure we switch back to main window even if there's an error
            if len(self.driver.window_handles) > 1:
                self.driver.close()
                self.driver.switch_to.window(handles[0])
            raise

    def process_payment_detail(self, excel_path):
        """Process payment detail report"""
        try:
            # Extract the payment detail table
            df = self.extract_payment_table_data(table_index=1, sheet_name="Payment Details")
            
            # If no data was found, return without creating sheet
            if df is None:
                logger.info("No payment data to process")
                return None
            
            # Only create Excel sheet if we have data
            if not df.empty:
                with pd.ExcelWriter(str(excel_path), engine='openpyxl', mode='a') as writer:
                    sheet_name = "Payment Details"
                    
                    # Remove sheet if it exists
                    if sheet_name in writer.book.sheetnames:
                        idx = writer.book.sheetnames.index(sheet_name)
                        writer.book.remove(writer.book.worksheets[idx])
                    
                    df.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False
                    )
                    
                    logger.info(f"Successfully exported {len(df)} payment details to sheet in {excel_path}")
            else:
                logger.info("No payment data to export")
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to process payment detail: {str(e)}")
            self.save_screenshot("payment_detail_error")
            raise

# Custom exception for security-related errors
class SecurityError(Exception):
    pass